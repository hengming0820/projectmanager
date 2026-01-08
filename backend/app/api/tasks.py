from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from typing import List, Optional
from datetime import datetime
from app.utils.datetime_utils import utc_now
from app.database import get_db
from app.schemas.task import TaskCreate, TaskUpdate, TaskResponse, TaskSubmit, TaskReview, TaskSkip, TaskSkipRequest, TaskSkipReview
from app.models.task import Task, TaskAttachment
from app.utils.file_utils import file_service
from app.models.project import Project
from app.models.user import User
from app.utils.security import get_current_user, get_current_admin_user
from app.utils.permissions import require_permission
from app.services.performance_service import performance_service
from app.services.notification_ws import manager as ws_manager
from app.services.stats_cache_service import stats_cache_service
import io
import csv
try:
    import openpyxl  # xlsx
except Exception:
    openpyxl = None
try:
    import xlrd  # xls
except Exception:
    xlrd = None
from app.utils.audit_logger import audit_logger
import logging
from app.services.cache_service import cache_service

# 配置日志
logger = logging.getLogger(__name__)

router = APIRouter()

# ========== 工具函数 ==========

def update_project_stats(db: Session, project_id: str):
    """
    更新项目的统计字段
    - total_tasks: 项目总任务数
    - assigned_tasks: 已分配任务数
    - completed_tasks: 已完成任务数（approved状态）
    """
    try:
        logger.info(f"📊 [ProjectStats] 开始更新项目统计: {project_id}")
        
        # 查询项目
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            logger.warning(f"⚠️ [ProjectStats] 项目不存在: {project_id}")
            return
        
        # 查询该项目的所有任务
        tasks = db.query(Task).filter(Task.project_id == project_id).all()
        
        # 统计
        total = len(tasks)
        assigned = len([t for t in tasks if t.assigned_to is not None])
        completed = len([t for t in tasks if t.status == 'approved'])
        
        # 更新项目字段
        project.total_tasks = total
        project.assigned_tasks = assigned
        project.completed_tasks = completed
        
        logger.info(f"✅ [ProjectStats] 项目统计已更新: {project_id} | 总数:{total} 已分配:{assigned} 已完成:{completed}")
        
        # 不在这里commit，让调用者决定何时commit
        
    except Exception as e:
        logger.error(f"❌ [ProjectStats] 更新项目统计失败: {project_id} | 错误: {e}")
        raise

@router.post("/init-test-data")
def init_test_data(
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("TaskPool"))
):
    """初始化测试数据（仅管理员，临时用于测试）"""
    logger.info(f"📊 [TaskAPI] 管理员初始化测试数据: {current_user.username}")
    
    # 检查是否已有任务数据
    existing_tasks = db.query(Task).count()
    if existing_tasks > 0:
        logger.info(f"📊 [TaskAPI] 已存在 {existing_tasks} 个任务，跳过初始化")
        return {
            "message": f"已存在 {existing_tasks} 个任务，无需初始化",
            "existing_tasks": existing_tasks
        }
    
    # 创建测试任务
    test_tasks = [
        Task(
            id="task5",
            title="膀胱CT标注任务005",
            description="标注膀胱CT影像中的肿瘤区域",
            project_id="proj1",
            status="pending",
            priority="high",
            created_by=current_user.id,
            image_url="/api/images/bladder005.jpg",
            score=55
        ),
        Task(
            id="task6",
            title="输尿管CT标注任务006",
            description="标注输尿管CT影像中的结石区域",
            project_id="proj1",
            status="pending",
            priority="medium",
            created_by=current_user.id,
            image_url="/api/images/ureter006.jpg",
            score=40
        ),
        Task(
            id="task7",
            title="肾脏CT标注任务007",
            description="标注左肾CT影像中的感染区域",
            project_id="proj1",
            status="pending",
            priority="low",
            created_by=current_user.id,
            image_url="/api/images/kidney007.jpg",
            score=35
        )
    ]
    
    for task in test_tasks:
        db.add(task)
    
    db.commit()
    
    logger.info(f"✅ [TaskAPI] 测试数据初始化完成: {len(test_tasks)} 个任务")
    return {
        "message": "测试数据初始化成功",
        "created_tasks": len(test_tasks),
        "task_ids": [task.id for task in test_tasks]
    }

@router.post("/", response_model=TaskResponse)
def create_task(
    task_data: TaskCreate,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("TaskPool"))
):
    """创建任务（仅管理员）"""
    # 验证项目是否存在
    project = db.query(Project).filter(Project.id == task_data.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    # 创建任务时初始化时间轴
    created_time = utc_now()
    initial_timeline = [{
        "type": "created",
        "time": created_time.isoformat(),
        "user_id": current_user.id,
        "user_name": getattr(current_user, 'real_name', None) or current_user.username
    }]

    # 填充创建者姓名
    creator_name = current_user.real_name if hasattr(current_user, 'real_name') else getattr(current_user, 'username', None)
    db_task = Task(
        **task_data.dict(),
        created_by=current_user.id,
        created_by_name=creator_name,
        timeline=initial_timeline
    )
    db.add(db_task)
    
    # ✅ 更新项目统计字段
    if task_data.project_id:
        try:
            update_project_stats(db, task_data.project_id)
        except Exception as e:
            logger.error(f"❌ [TaskAPI] 更新项目统计失败: {e}")
    
    db.commit()
    db.refresh(db_task)

    # ✅ 清除缓存（三重清除策略）
    # 1. 清除项目的所有任务缓存
    cache_service.invalidate_tasks_cache(task_data.project_id)
    # 2. 清除跨项目的任务缓存（任务池可能查看所有项目）
    cache_service.invalidate_tasks_cache()
    # 3. 清除项目详情
    cache_service.invalidate_project_detail_cache(task_data.project_id)
    # 清除统计缓存
    stats_cache_service.invalidate_dashboard_stats()
    stats_cache_service.invalidate_project_stats(task_data.project_id)

    logger.info(f"✅ [TaskAPI] 任务创建成功: {db_task.id}, 时间轴已初始化: {len(db_task.timeline or [])} 个事件")
    logger.info(f"✅ [TaskAPI] 任务创建缓存已清除: project={task_data.project_id}, 所有视图已刷新")
    return db_task

@router.get("/", include_in_schema=True)
def get_tasks(
    project_id: Optional[str] = None,
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    include_completed_projects: bool = False,  # ✅ 新增：是否包含完结项目的任务
    db: Session = Depends(get_db),
    current_user = Depends(require_permission(["TaskPool", "ProjectManagement"]))  # 允许TaskPool或ProjectManagement权限
):
    """获取任务列表（带Redis缓存）"""
    logger.info(f"📋 [TaskAPI] 获取任务列表 - 用户: {current_user.username}, 角色: {current_user.role}, ID: {current_user.id}")
    logger.info(f"📋 [TaskAPI] 查询参数 - project_id: {project_id}, status: {status}, assigned_to: {assigned_to}, skip: {skip}, limit: {limit}, include_completed_projects: {include_completed_projects}")
    
    # 生成缓存key
    cache_key = f"tasks:list:{project_id or 'all'}:{status or 'all'}:{assigned_to or 'all'}:{skip}:{limit}:{include_completed_projects}"
    
    # 尝试从缓存获取
    cached_data = cache_service.get(cache_key)
    if cached_data:
        logger.info(f"🎯 任务列表缓存命中: {cache_key}")
        return cached_data
    
    # 使用join查询以包含项目信息
    from app.models.project import Project
    query = db.query(Task).join(Project, Task.project_id == Project.id)
    
    # ✅ 根据参数决定是否过滤完结项目的任务
    if not include_completed_projects:
        # 默认情况：过滤掉完结项目的任务（用于任务池、工作台等）
        query = query.filter(Project.status != "completed")
        logger.info(f"🔒 [TaskAPI] 已过滤完结项目的任务")
    else:
        # 包含完结项目的任务（用于项目详情、历史查看等）
        logger.info(f"📖 [TaskAPI] 包含完结项目的任务（历史查看模式）")
    
    if project_id:
        query = query.filter(Task.project_id == project_id)
    if status:
        # 后端聚合筛选：accepted 表示已被接收流转的任务集合
        if status == "accepted":
            query = query.filter(Task.status.in_(["submitted", "skip_pending", "skipped", "approved", "rejected"]))
        else:
            query = query.filter(Task.status == status)
    if assigned_to:
        query = query.filter(Task.assigned_to == assigned_to)
    
    # 统一权限管理：通过菜单权限控制，不再硬编码角色检查
    # 如果用户能访问 TaskPool，就能查看所有任务数据
    logger.info(f"🔐 [TaskAPI] 统一权限管理 - 用户: {current_user.username}, 角色: {current_user.role}")
    logger.info(f"✅ [TaskAPI] 用户已通过 TaskPool 权限验证，可查看所有任务数据")
    
    # 执行查询前记录SQL
    total_tasks = query.count()
    logger.info(f"📊 [TaskAPI] 权限过滤后的任务总数: {total_tasks}")
    
    tasks = query.offset(skip).limit(limit).all()
    logger.info(f"✅ [TaskAPI] 返回任务数量: {len(tasks)} / 总数: {total_tasks}")
    
    # 为每个任务设置项目名称
    task_responses = []
    for task in tasks:
        task_dict = {
            "id": task.id,
            "title": task.title,
            "description": task.description,
            "project_id": task.project_id,
            "project_name": task.project.name if task.project else "未知项目",
            "status": task.status,
            "priority": task.priority,
            "assigned_to": task.assigned_to,
            "assigned_to_name": getattr(task, 'assigned_to_name', None),
            "created_by": task.created_by,
            "created_by_name": getattr(task, 'created_by_name', None),
            "image_url": task.image_url,
            "annotation_data": task.annotation_data,
            "score": task.score,
            "assigned_at": task.assigned_at,
            "submitted_at": task.submitted_at,
            "reviewed_by": task.reviewed_by,
            "reviewed_by_name": getattr(task, 'reviewed_by_name', None),
            "reviewed_at": task.reviewed_at,
            "review_comment": task.review_comment,
            "created_at": task.created_at,
            "updated_at": task.updated_at,
            "attachments": task.attachments or [],
            "timeline": task.timeline or []
        }
        task_responses.append(task_dict)
    
    # 记录返回的任务详情（仅前3个）
    for i, task_dict in enumerate(task_responses[:3]):
        logger.info(f"📄 [TaskAPI] 任务 {i+1}: ID={task_dict['id']}, 标题={task_dict['title']}, 项目={task_dict['project_name']}, 创建者={task_dict['created_by']}, 分配给={task_dict['assigned_to']}, 状态={task_dict['status']}")
    
    # 构建响应
    result = {"list": task_responses, "total": total_tasks}
    
    # 写入缓存（5分钟）
    cache_service.set(cache_key, result, expire=300)
    logger.info(f"💾 任务列表写入缓存: {cache_key}")
    
    # 返回分页结构
    return result

@router.get("/{task_id}", response_model=TaskResponse)
def get_task(
    task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("TaskPool"))
):
    """获取任务详情（带Redis缓存）"""
    logger.info(f"🔍 [TaskAPI] 获取任务详情: task_id={task_id}, user={current_user.username}")

    # 暂时禁用任务详情缓存（ORM对象序列化问题）
    # cache_key = f"tasks:detail:{task_id}"
    # cached_task = cache_service.get(cache_key)
    # if cached_task:
    #     logger.info(f"🎯 任务详情缓存命中: {task_id}")
    #     return cached_task
    
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        logger.warning(f"❌ [TaskAPI] 任务不存在: {task_id}")
        # 查询所有任务ID用于调试
        all_task_ids = db.query(Task.id).all()
        logger.info(f"📋 [TaskAPI] 数据库中的所有任务ID: {[t.id for t in all_task_ids]}")
        raise HTTPException(status_code=404, detail="任务不存在")

    # 统一权限管理：通过菜单权限控制，不再硬编码角色检查
    # 如果用户能访问 TaskPool，就能查看任务详情
    logger.info(f"✅ [TaskAPI] 用户已通过 TaskPool 权限验证，可查看任务详情: {task_id}")

    logger.info(f"✅ [TaskAPI] 任务详情获取成功: {task_id}, timeline事件数: {len(task.timeline or [])}")
    
    # 暂时禁用单个任务详情的缓存，因为ORM对象序列化问题
    # 任务列表已经使用字典格式，不受影响
    # cache_service.set(cache_key, task, expire=300)
    # logger.debug(f"💾 任务详情写入缓存: {task_id}")
    
    return task

@router.post("/{task_id}/claim")
def claim_task(
    task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("TaskPool"))
):
    """领取任务"""
    logger.info(f"📦 [TaskAPI] 用户领取任务: {current_user.username} -> {task_id}")
    
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        logger.warning(f"❌ [TaskAPI] 任务不存在: {task_id}")
        raise HTTPException(status_code=404, detail="任务不存在")
    
    if task.status != "pending":
        logger.warning(f"❌ [TaskAPI] 任务状态不允许领取: {task_id}, 当前状态: {task.status}")
        raise HTTPException(status_code=400, detail="任务状态不允许领取")
    
    if task.assigned_to:
        logger.warning(f"❌ [TaskAPI] 任务已被领取: {task_id}, 领取人: {task.assigned_to}")
        raise HTTPException(status_code=400, detail="任务已被领取")
    
    # 领取前检查用户当前活跃任务数量（in_progress + submitted + rejected 不超过 3）
    active_q = db.query(Task).filter(
        Task.assigned_to == current_user.id,
        Task.status.in_(["in_progress", "submitted", "rejected"])
    )
    active_count = active_q.count()
    logger.info(f"📏 [TaskAPI] 当前活跃任务数(仅本用户): {active_count} (用户: {current_user.id}, 角色: {current_user.role})")
    if active_count >= 3:
        # 记录详细任务用于排查
        details = [
            {
                "id": t.id,
                "status": t.status,
                "title": t.title,
                "project_id": t.project_id,
            }
            for t in active_q.limit(10).all()
        ]
        logger.info(f"📄 [TaskAPI] 活跃任务明细(最多10条): {details}")
    # 上限规则：仅对非管理员生效
    if (current_user.role or '').lower() != 'admin' and active_count >= 3:
        logger.warning(f"❌ [TaskAPI] 已达到可领取任务上限(3): 用户 {current_user.id}")
        raise HTTPException(status_code=400, detail="可领取任务数量已达上限(3)")

    # 更新任务状态
    task.status = "in_progress"
    task.assigned_to = current_user.id
    task.assigned_to_name = getattr(current_user, 'real_name', None) or getattr(current_user, 'username', None)
    assigned_time = utc_now()
    task.assigned_at = assigned_time
    # 记录时间轴
    events = list(task.timeline or [])  # 创建新列表，避免SQLAlchemy的可变对象问题
    events.append({
        "type": "claimed",
        "time": assigned_time.isoformat(),
        "user_id": current_user.id,
        "user_name": getattr(current_user, 'real_name', None) or current_user.username
    })
    task.timeline = events
    flag_modified(task, 'timeline')  # 明确告诉SQLAlchemy字段已修改
    
    # ✅ 更新项目统计字段
    if task.project_id:
        try:
            update_project_stats(db, task.project_id)
        except Exception as e:
            logger.error(f"❌ [TaskAPI] 更新项目统计失败: {e}")
    
    db.commit()
    
    # ✅ 清除缓存（重要：三重清除策略）
    # 1. 清除领取者的任务缓存
    cache_service.invalidate_tasks_cache(task.project_id, current_user.id)
    # 2. 清除项目的所有任务缓存
    cache_service.invalidate_tasks_cache(task.project_id)
    # 3. 清除跨项目的任务缓存（任务池可能查看所有项目）
    cache_service.invalidate_tasks_cache()
    # 4. 清除任务详情
    cache_service.invalidate_task_detail_cache(task_id)
    cache_service.invalidate_project_detail_cache(task.project_id)
    
    logger.info(f"✅ [TaskAPI] 任务领取成功: {task_id} -> 用户 {current_user.username}")
    logger.info(f"✅ [TaskAPI] 任务领取缓存已清除: project={task.project_id}, user={current_user.id}, 所有视图已刷新")
    audit_logger.info(f"user_id={current_user.id} action=claim_task task_id={task_id}")
    return {
        "success": True,
        "message": "任务领取成功",
        "task_id": task_id,
        "status": "in_progress",
        "assigned_to": current_user.id
    }

@router.post("/{task_id}/submit")
async def submit_task(
    task_id: str,
    task_submit: TaskSubmit,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """提交任务审核"""
    logger.info(f"📝 [TaskAPI] 用户提交任务: {current_user.username} -> {task_id}")
    
    # 详细记录接收到的数据
    logger.info(f"📋 [TaskAPI] TaskSubmit数据: annotation_data={task_submit.annotation_data}, comment={task_submit.comment}, organ_count={task_submit.organ_count}")
    
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        logger.warning(f"❌ [TaskAPI] 任务不存在: {task_id}")
        raise HTTPException(status_code=404, detail="任务不存在")
    
    logger.info(f"📊 [TaskAPI] 任务当前状态: id={task.id}, status={task.status}, assigned_to={task.assigned_to}, current_user={current_user.id}")
    
    if task.assigned_to != current_user.id:
        logger.warning(f"❌ [TaskAPI] 只能提交自己领取的任务: {task_id}, 领取人: {task.assigned_to}, 当前用户: {current_user.id}")
        raise HTTPException(status_code=403, detail="只能提交自己领取的任务")
    
    # 允许 in_progress 和 rejected 状态的任务提交
    if task.status not in ["in_progress", "rejected"]:
        logger.warning(f"❌ [TaskAPI] 任务状态不允许提交: {task_id}, 当前状态: {task.status}")
        raise HTTPException(status_code=400, detail="只有进行中或已驳回的任务才能提交")
    
    # 更新任务状态
    task.status = "submitted"
    task.annotation_data = task_submit.annotation_data
    submitted_time = utc_now()
    task.submitted_at = submitted_time
    # 时间轴
    events = list(task.timeline or [])  # 创建新列表，避免SQLAlchemy的可变对象问题
    events.append({
        "type": "submitted",
        "time": submitted_time.isoformat(),
        "user_id": current_user.id,
        "user_name": getattr(current_user, 'real_name', None) or current_user.username,
        "comment": task_submit.comment or "",
        "organ_count": getattr(task_submit, 'organ_count', None)
    })
    task.timeline = events
    flag_modified(task, 'timeline')  # 明确告诉SQLAlchemy字段已修改
    
    db.commit()
    
    # ✅ 清除缓存（重要：三重清除策略）
    # 1. 清除提交者的任务缓存
    cache_service.invalidate_tasks_cache(task.project_id, current_user.id)
    # 2. 清除项目的所有任务缓存
    cache_service.invalidate_tasks_cache(task.project_id)
    # 3. 清除跨项目的任务缓存（审核员可能查看所有项目）
    cache_service.invalidate_tasks_cache()
    # 4. 清除任务详情
    cache_service.invalidate_task_detail_cache(task_id)
    # 清除统计缓存
    stats_cache_service.invalidate_performance_stats(current_user.id)
    stats_cache_service.invalidate_dashboard_stats()
    stats_cache_service.invalidate_project_stats(task.project_id)
    
    logger.info(f"✅ [TaskAPI] 任务提交缓存已清除: project={task.project_id}, user={current_user.id}, 所有视图已刷新")
    
    logger.info(f"✅ [TaskAPI] 任务提交成功: {task_id} -> 等待审核")
    audit_logger.info(f"user_id={current_user.id} action=submit_task task_id={task_id}")
    
    # ✅ Redis Pub/Sub 通知审核员
    try:
        # 统计待审核任务数量
        pending_count = db.query(Task).filter(Task.status == "submitted").count()
        logger.info(f"🔔 [TaskAPI] 通知审核员: 新任务待审核，待审核总数: {pending_count}")
        
        # 统一的通知消息
        notification_message = {
            "type": "task_submitted",
            "title": "新任务待审核",
            "content": f"{current_user.real_name or current_user.username} 提交了任务《{task.title}》",
            "pending": pending_count,
            "task_id": task_id
        }
        
        # 通知审核员（WebSocket 管理器内部会使用 Redis Pub/Sub）
        logger.info(f"🔔 [TaskAPI] 向审核员广播任务提交通知")
        await ws_manager.broadcast_to_role(role="reviewer", message=notification_message)
        
        # 通知管理员
        logger.info(f"🔔 [TaskAPI] 向管理员广播任务提交通知")
        await ws_manager.broadcast_to_role(role="admin", message=notification_message)
    except Exception as _e:
        logger.warning(f"通知审核员失败: {_e}")
    return {
        "success": True,
        "message": "任务提交成功",
        "task_id": task_id,
        "status": "submitted"
    }

@router.post("/{task_id}/upload-annotation-images")
async def upload_annotation_images(
    task_id: str,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """上传标注截图（MinIO），返回URL列表并写入附件表"""
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    urls = await file_service.upload_annotation_screenshots(task_id, files)
    for url in urls:
        db.add(TaskAttachment(
            task_id=task_id,
            file_name=url.split('/')[-1],
            file_url=url,
            file_type='image',
            attachment_type='annotation_screenshot',
            uploaded_by=current_user.id
        ))
    db.commit()
    
    # ✅ 清除缓存
    cache_service.invalidate_task_detail_cache(task_id)
    logger.info(f"✅ [TaskAPI] 上传附件成功,任务详情缓存已清除: task={task_id}")
    
    return {"urls": urls}

@router.post("/import")
async def import_tasks(
    file: UploadFile = File(...),
    project_id: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("TaskPool"))
):
    """批量导入任务，支持 .xlsx/.xls/.csv
    期望列：title, description, priority, image_url, estimated_hours，可选 assigned_to, project_id
    如果未提供 project_id 字段，则使用表单中的 project_id
    """
    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="仅支持xlsx/xls/csv 文件")

    try:
        content = await file.read()
        rows = []

        if filename.endswith(".csv"):
            # 尝试多种编码（Windows下Excel常用GBK）
            decoded_text: str
            try:
                decoded_text = content.decode("utf-8-sig")
            except Exception:
                try:
                    decoded_text = content.decode("gbk")
                except Exception:
                    decoded_text = content.decode("latin1")
            text_stream = io.StringIO(decoded_text)
            reader = csv.DictReader(text_stream)
            for row in reader:
                rows.append({(k or '').strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()})
        elif filename.endswith(".xlsx") and openpyxl:
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            headers = [str(c.value).strip() if c.value is not None else '' for c in header_cells]
            for r in ws.iter_rows(min_row=2, values_only=True):
                row = {headers[i]: (str(r[i]).strip() if r[i] is not None else '') for i in range(len(headers))}
                rows.append(row)
        elif filename.endswith(".xls") and xlrd:
            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_index(0)
            headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
            for r in range(1, sheet.nrows):
                row = {headers[c]: str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)}
                rows.append(row)
        else:
            raise HTTPException(status_code=500, detail="服务器缺少Excel解析依赖，请安装openpyxl/xlrd")

        created = 0
        errors = []

        def pick(row: dict, names):
            for n in names:
                if n in row and str(row[n]).strip() != '':
                    return str(row[n]).strip()
            lower = {k.lower(): v for k, v in row.items()}
            for n in names:
                if n.lower() in lower and str(lower[n.lower()]).strip() != '':
                    return str(lower[n.lower()]).strip()
            return None

        for idx, row in enumerate(rows, start=2):
            try:
                title = pick(row, ["title", "任务标题", "名称"]) or ''
                if not title:
                    errors.append(f"第{idx}行缺少任务标题")
                    continue
                description = pick(row, ["description", "任务描述"]) or ''
                priority = pick(row, ["priority", "优先级"]) or 'medium'
                # 兼容中文优先级
                priority_map = {
                    '低': 'low', 'low': 'low',
                    '中': 'medium', '中等': 'medium', 'medium': 'medium',
                    '高': 'high', 'high': 'high',
                    '紧急': 'urgent', 'urgent': 'urgent'
                }
                priority = priority_map.get(priority, priority)
                image_url = pick(row, ["image_url", "影像URL", "影像链接", "图片链接"]) or None
                est_hours_raw = pick(row, ["estimated_hours", "预计工时"]) or '0'
                try:
                    estimated_hours = float(est_hours_raw)
                except Exception:
                    estimated_hours = 0.0
                assigned_to = pick(row, ["assigned_to", "标注员ID"]) or None

                project_id_value = project_id or pick(row, ["project_id", "项目ID"]) or None
                if not project_id_value:
                    errors.append(f"第{idx}行缺少项目ID（请在弹窗选择项目或在文件中提供project_id）")
                    continue

                project = db.query(Project).filter(Project.id == project_id_value).first()
                if not project:
                    errors.append(f"第{idx}行项目不存在: {project_id_value}")
                    continue

                task = Task(
                    title=title,
                    description=description,
                    project_id=project_id_value,
                    priority=priority if priority in ["low", "medium", "high", "urgent"] else "medium",
                    status='pending',
                    created_by=current_user.id,
                )
                if image_url:
                    task.image_url = image_url
                task.annotation_data = {"estimated_hours": estimated_hours}
                if assigned_to:
                    user = db.query(User).filter(User.id == assigned_to).first()
                    if user:
                        task.assigned_to = user.id

                db.add(task)
                created += 1
            except Exception as e:
                logging.exception("导入任务失败（行%r）", idx)
                errors.append(f"第{idx}行导入失败: {e}")

        # ✅ 更新项目统计字段
        if project_id:
            try:
                update_project_stats(db, project_id)
            except Exception as e:
                logger.error(f"❌ [TaskAPI] 批量导入后更新项目统计失败: {e}")

        db.commit()
        
        # ✅ 清除缓存（批量导入后）
        if project_id:
            # 1. 清除项目的所有任务缓存
            cache_service.invalidate_tasks_cache(project_id)
            # 2. 清除跨项目的任务缓存
            cache_service.invalidate_tasks_cache()
            # 3. 清除项目详情
            cache_service.invalidate_project_detail_cache(project_id)
            # 清除统计缓存
            stats_cache_service.invalidate_dashboard_stats()
            stats_cache_service.invalidate_project_stats(project_id)
            logger.info(f"✅ [TaskAPI] 批量导入{created}个任务后,缓存已清除: project={project_id}, 所有视图已刷新")
        
        return {"success": True, "created": created, "failed": len(errors), "errors": errors, "message": f"成功导入{created}条，失败{len(errors)}条"}
    except HTTPException:
        raise
    except Exception as e:
        logging.exception("导入任务异常")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{task_id}/upload-review-images")
async def upload_review_images(
    task_id: str,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(require_permission(["TaskReview", "TeamPerformance"]))
):
    """上传审核打回截图（MinIO），返回URL列表并写入附件表"""
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    urls = await file_service.upload_review_screenshots(task_id, files)
    for url in urls:
        db.add(TaskAttachment(
            task_id=task_id,
            file_name=url.split('/')[-1],
            file_url=url,
            file_type='image',
            attachment_type='review_screenshot',
            uploaded_by=current_user.id
        ))
    db.commit()
    
    # ✅ 清除缓存
    cache_service.invalidate_task_detail_cache(task_id)
    logger.info(f"✅ [TaskAPI] 上传附件成功,任务详情缓存已清除: task={task_id}")
    
    return {"urls": urls}

@router.post("/{task_id}/upload-skip-images")
async def upload_skip_images(
    task_id: str,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("TaskPool"))
):
    """上传跳过原因截图（MinIO），返回URL列表并写入附件表"""
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    # 直接复用review截图上传实现
    urls = await file_service.upload_review_screenshots(task_id, files)
    for url in urls:
        db.add(TaskAttachment(
            task_id=task_id,
            file_name=url.split('/')[-1],
            file_url=url,
            file_type='image',
            attachment_type='skip_screenshot',
            uploaded_by=current_user.id
        ))
    db.commit()
    
    # ✅ 清除缓存
    cache_service.invalidate_task_detail_cache(task_id)
    logger.info(f"✅ [TaskAPI] 上传附件成功,任务详情缓存已清除: task={task_id}")
    
    return {"urls": urls}

@router.post("/{task_id}/abandon")
def abandon_task(
    task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("MyWorkspace"))
):
    """放弃任务"""
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    if task.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="只能放弃自己领取的任务")
    
    if task.status != "in_progress":
        raise HTTPException(status_code=400, detail="任务状态不允许放弃")
    
    task.status = "pending"
    task.assigned_to = None
    task.assigned_to_name = None
    task.assigned_at = None
    
    # ✅ 更新项目统计字段
    if task.project_id:
        try:
            update_project_stats(db, task.project_id)
        except Exception as e:
            logger.error(f"❌ [TaskAPI] 更新项目统计失败: {e}")
    
    db.commit()
    
    # ✅ 清除缓存（重要：三重清除策略）
    # 1. 清除放弃者的任务缓存
    cache_service.invalidate_tasks_cache(task.project_id, current_user.id)
    # 2. 清除项目的所有任务缓存
    cache_service.invalidate_tasks_cache(task.project_id)
    # 3. 清除跨项目的任务缓存（任务池可能查看所有项目）
    cache_service.invalidate_tasks_cache()
    # 4. 清除任务详情
    cache_service.invalidate_task_detail_cache(task_id)
    cache_service.invalidate_project_detail_cache(task.project_id)
    
    logger.info(f"✅ [TaskAPI] 任务放弃缓存已清除: project={task.project_id}, user={current_user.id}, 所有视图已刷新")
    
    return {
        "success": True,
        "message": "任务放弃成功",
        "task_id": task_id,
        "status": "pending"
    }

@router.post("/{task_id}/restart")
def restart_task(
    task_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("MyWorkspace"))
):
    """重新开始驳回的任务"""
    logger.info(f"🔄 [TaskAPI] 用户重新开始驳回任务: {current_user.username} -> {task_id}")
    
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        logger.warning(f"❌ [TaskAPI] 任务不存在: {task_id}")
        raise HTTPException(status_code=404, detail="任务不存在")
    
    if task.assigned_to != current_user.id:
        logger.warning(f"❌ [TaskAPI] 只能重新开始自己被分配的任务: {task_id}, 分配给: {task.assigned_to}, 当前用户: {current_user.id}")
        raise HTTPException(status_code=403, detail="只能重新开始自己被分配的任务")
    
    if task.status != "rejected":
        logger.warning(f"❌ [TaskAPI] 只有已驳回的任务才能重新开始: {task_id}, 当前状态: {task.status}")
        raise HTTPException(status_code=400, detail="只有已驳回的任务才能重新开始")
    
    # 更新任务状态，重新开始标注
    task.status = "in_progress"
    task.submitted_at = None  # 清除之前的提交时间
    task.reviewed_at = None   # 清除审核时间
    task.review_comment = None  # 清除审核评论
    # 时间轴
    events = list(task.timeline or [])  # 创建新列表，避免SQLAlchemy的可变对象问题
    events.append({
        "type": "restarted",
        "time": utc_now().isoformat(),
        "user_id": current_user.id,
        "user_name": current_user.username
    })
    task.timeline = events
    flag_modified(task, 'timeline')  # 明确告诉SQLAlchemy字段已修改
    
    db.commit()
    
    # ✅ 清除缓存（重要：三重清除策略）
    # 1. 清除重启者的任务缓存
    cache_service.invalidate_tasks_cache(task.project_id, current_user.id)
    # 2. 清除项目的所有任务缓存
    cache_service.invalidate_tasks_cache(task.project_id)
    # 3. 清除跨项目的任务缓存（审核页面可能查看所有项目）
    cache_service.invalidate_tasks_cache()
    # 4. 清除任务详情
    cache_service.invalidate_task_detail_cache(task_id)
    
    logger.info(f"✅ [TaskAPI] 驳回任务重新开始成功: {task_id} -> 状态: in_progress")
    logger.info(f"✅ [TaskAPI] 任务重启缓存已清除: project={task.project_id}, user={current_user.id}, 所有视图已刷新")
    audit_logger.info(f"user_id={current_user.id} action=restart_task task_id={task_id}")
    return {
        "success": True,
        "message": "任务已重新开始，可以进行标注",
        "task_id": task_id,
        "status": "in_progress",
        "assigned_to": current_user.id
    }

@router.post("/{task_id}/review")
async def review_task(
    task_id: str,
    task_review: TaskReview,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("TaskReview"))
):
    """审核任务（需菜单权限 TaskReview）"""
    logger.info(f"📋 [TaskAPI] 开始审核任务: {task_id}, 审核人: {current_user.username}")
    
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        logger.warning(f"❌ [TaskAPI] 任务不存在: {task_id}")
        raise HTTPException(status_code=404, detail="任务不存在")
    
    if task.status != "submitted":
        logger.warning(f"❌ [TaskAPI] 任务状态不允许审核: {task_id}, 当前状态: {task.status}")
        raise HTTPException(status_code=400, detail="任务状态不允许审核")
    
    # 验证审核意见：驳回时必须填写意见
    if task_review.action == "reject" and not (task_review.comment and task_review.comment.strip()):
        logger.warning(f"❌ [TaskAPI] 驳回任务时必须填写审核意见: {task_id}")
        raise HTTPException(status_code=400, detail="驳回任务时必须填写审核意见")
    
    # 记录原始状态
    original_status = task.status
    assigned_user_id = task.assigned_to
    
    # 更新任务状态
    task.status = "approved" if task_review.action == "approve" else "rejected"
    task.reviewed_by = current_user.id
    task.reviewed_by_name = getattr(current_user, 'real_name', None) or getattr(current_user, 'username', None)
    reviewed_time = utc_now()
    task.reviewed_at = reviewed_time
    task.review_comment = task_review.comment or ""  # 确保不会保存None
    # 时间轴
    events = list(task.timeline or [])  # 创建新列表，避免SQLAlchemy的可变对象问题
    events.append({
        "type": "reviewed",
        "time": reviewed_time.isoformat(),
        "user_id": current_user.id,
        "user_name": getattr(current_user, 'real_name', None) or current_user.username,
        "action": task_review.action,
        "comment": task_review.comment or "",
        "score": getattr(task_review, 'score', None)
    })
    task.timeline = events
    flag_modified(task, 'timeline')  # 明确告诉SQLAlchemy字段已修改

    # 如果是打回并携带了截图URL，保存为附件记录
    if task_review.action != "approve" and getattr(task_review, 'reject_images', None):
        for url in (task_review.reject_images or [])[:10]:
            attach = TaskAttachment(
                task_id=task.id,
                file_name=url.split('/')[-1],
                file_url=url,
                file_type='image',
                attachment_type='review_screenshot',
                uploaded_by=current_user.id
            )
            db.add(attach)
    
    # 如果提供了评分，更新任务评分
    if hasattr(task_review, 'score') and task_review.score is not None:
        task.score = task_review.score
    
    logger.info(f"✅ [TaskAPI] 任务状态更新: {task_id} {original_status} -> {task.status}")
    
    # ✅ 更新项目统计字段
    if task.project_id:
        try:
            update_project_stats(db, task.project_id)
        except Exception as e:
            logger.error(f"❌ [TaskAPI] 更新项目统计失败: {e}")
            # 不影响任务审核流程
    
    db.commit()
    
    # ✅ 清除缓存（重要：三重清除策略）
    # 1. 清除标注员的任务缓存
    cache_service.invalidate_tasks_cache(task.project_id, assigned_user_id)
    # 2. 清除项目的所有任务缓存
    cache_service.invalidate_tasks_cache(task.project_id)
    # 3. 清除跨项目的任务缓存（审核员可能查看所有项目）
    cache_service.invalidate_tasks_cache()
    # 4. 清除任务详情和项目详情
    cache_service.invalidate_task_detail_cache(task_id)
    cache_service.invalidate_project_detail_cache(task.project_id)
    # 清除统计缓存
    stats_cache_service.invalidate_performance_stats(assigned_user_id)
    stats_cache_service.invalidate_performance_stats(current_user.id)
    stats_cache_service.invalidate_dashboard_stats()
    stats_cache_service.invalidate_project_stats(task.project_id)
    
    logger.info(f"✅ [TaskAPI] 任务审核缓存已清除: project={task.project_id}, user={assigned_user_id}, 所有视图已刷新")
    
    # ✅ 通知标注员（仅发送一次，使用统一的通知类型）
    # 注意：移除了重复的通知发送代码，现在只在下方统一发送
    
    # 审核通过后，立即为用户增加绩效分数
    performance_message = ""
    if task_review.action == "approve" and assigned_user_id:
        try:
            logger.info(f"📈 [TaskAPI] 开始更新用户绩效: {assigned_user_id}")
            
            # 获取用户信息
            assigned_user = db.query(User).filter(User.id == assigned_user_id).first()
            user_name = assigned_user.username if assigned_user else f"用户ID:{assigned_user_id}"
            
            # 使用新的绩效增加方法
            score_added = task_review.score if task_review.score else 1  # 默认1分
            performance_data = performance_service.add_performance_score(
                db=db, 
                user_id=assigned_user_id, 
                task_score=score_added,
                period="monthly"
            )
            
            performance_message = f"已为 {user_name} 增加了 {score_added} 点绩效，总分: {performance_data['total_score']}"
            
            logger.info(f"✅ [TaskAPI] 绩效更新成功: 用户 {user_name}({assigned_user_id}), 本次增加: {score_added}分, 总分: {performance_data['total_score']}")
        except Exception as e:
            logger.error(f"❌ [TaskAPI] 绩效更新失败: {e}")
            performance_message = f"绩效更新失败: {str(e)}"
            # 绩效更新失败不影响任务审核结果
    
    message = "任务审核通过" if task_review.action == "approve" else "任务已打回重标"
    
    # 如果有绩效消息，追加到主消息中
    if performance_message:
        message = f"{message}，{performance_message}"
    
    logger.info(f"✅ [TaskAPI] 任务审核完成: {task_id}, 结果: {message}")
    audit_logger.info(f"user_id={current_user.id} action=review_task task_id={task_id} result={task.status}")

    # ✅ 通知标注员（统一通知发送点，避免重复）
    try:
        if assigned_user_id:
            action_text = "通过" if task_review.action == "approve" else "驳回"
            notification_type = "task_approved" if task_review.action == "approve" else "task_rejected"
            content = (
                f"恭喜你，你所提交的任务《{task.title}》审核通过！" if task_review.action == "approve"
                else f"你的任务《{task.title}》需修订，请修改"
            )
            
            payload = {
                "type": notification_type,
                "title": f"任务审核结果：{action_text}",
                "content": content,
                "task_id": task_id,
                "action": task_review.action,
                "comment": task_review.comment or "",
                "reviewer": current_user.real_name or current_user.username
            }
            
            logger.info(f"🔔 [TaskAPI] 向标注员 {assigned_user_id} 发送审核通知: {notification_type} - {task.title}")
            
            # 使用 WebSocket 管理器发送（内部自动使用 Redis Pub/Sub）
            await ws_manager.send_to_user_id(assigned_user_id, payload)
            
            logger.info(f"✅ [TaskAPI] 审核通知已发送: 用户 {assigned_user_id}, 结果: {action_text}")
    except Exception as e:
        logger.error(f"❌ [TaskAPI] 发送审核通知失败: {e}")
    
    return {
        "success": True,
        "message": message,
        "task_id": task_id,
        "status": task.status,
        "assigned_to": assigned_user_id
    }

@router.post("/{task_id}/skip")
def skip_task(
    task_id: str,
    payload: TaskSkip,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("TaskPool"))
):
    """跳过任务（软删除，需菜单权限 TaskPool）：记录原因与截图，状态置为 skipped"""
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    # 已完成/已通过不允许跳过
    if task.status in ["approved"]:
        raise HTTPException(status_code=400, detail="已完成任务不允许跳过")

    # 更新状态与字段
    now = utc_now()
    task.status = "skipped"
    task.skipped_at = now
    task.skip_reason = payload.reason
    task.skip_images = payload.images or []
    # 时间轴事件
    events = list(task.timeline or [])
    events.append({
        "type": "skipped",
        "time": now.isoformat(),
        "user_id": current_user.id,
        "user_name": getattr(current_user, 'real_name', None) or current_user.username,
        "reason": payload.reason,
        "images": (payload.images or [])[:10]
    })
    task.timeline = events
    flag_modified(task, 'timeline')
    db.commit()
    
    # ✅ 清除缓存（三重清除策略）
    # 1. 清除项目的所有任务缓存
    cache_service.invalidate_tasks_cache(task.project_id)
    # 2. 清除跨项目的任务缓存
    cache_service.invalidate_tasks_cache()
    # 3. 清除任务详情
    cache_service.invalidate_task_detail_cache(task_id)
    logger.info(f"✅ [TaskAPI] 任务跳过缓存已清除: project={task.project_id}, 所有视图已刷新")
    
    return {"success": True, "message": "任务已标记为已跳过", "task_id": task_id, "status": "skipped"}

@router.post("/{task_id}/request-skip")
async def request_skip_task(
    task_id: str,
    payload: TaskSkipRequest,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """申请跳过任务"""
    logger.info(f"📋 [TaskAPI] 用户申请跳过任务: {current_user.username} -> {task_id}")
    
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        logger.warning(f"❌ [TaskAPI] 任务不存在: {task_id}")
        raise HTTPException(status_code=404, detail="任务不存在")
    
    # 验证权限：只能申请跳过自己领取的任务
    if task.assigned_to != current_user.id:
        logger.warning(f"❌ [TaskAPI] 只能申请跳过自己领取的任务: {task_id}, 领取人: {task.assigned_to}, 当前用户: {current_user.id}")
        raise HTTPException(status_code=403, detail="只能申请跳过自己领取的任务")
    
    # 验证状态：只有进行中的任务才能申请跳过
    if task.status != "in_progress":
        logger.warning(f"❌ [TaskAPI] 任务状态不允许申请跳过: {task_id}, 当前状态: {task.status}")
        raise HTTPException(status_code=400, detail="只有进行中的任务才能申请跳过")
    
    # 更新任务状态
    task.status = "skip_pending"
    requested_time = utc_now()
    task.skip_requested_at = requested_time
    task.skip_request_reason = payload.reason
    task.skip_request_images = payload.images or []
    task.skip_requested_by = current_user.id
    
    # 记录时间轴
    events = list(task.timeline or [])
    events.append({
        "type": "skip_requested",
        "time": requested_time.isoformat(),
        "user_id": current_user.id,
        "user_name": getattr(current_user, 'real_name', None) or current_user.username,
        "reason": payload.reason,
        "images": (payload.images or [])[:10]
    })
    task.timeline = events
    flag_modified(task, 'timeline')
    
    db.commit()
    
    # ✅ 清除缓存（重要：三重清除策略）
    # 1. 清除申请者的任务缓存
    cache_service.invalidate_tasks_cache(task.project_id, current_user.id)
    # 2. 清除项目的所有任务缓存
    cache_service.invalidate_tasks_cache(task.project_id)
    # 3. 清除跨项目的任务缓存（审核页面可能查看所有项目）
    cache_service.invalidate_tasks_cache()
    # 4. 清除任务详情
    cache_service.invalidate_task_detail_cache(task_id)
    
    logger.info(f"✅ [TaskAPI] 跳过申请提交成功: {task_id} -> 等待审核")
    logger.info(f"✅ [TaskAPI] 跳过申请缓存已清除: project={task.project_id}, user={current_user.id}, 所有视图已刷新")
    audit_logger.info(f"user_id={current_user.id} action=request_skip_task task_id={task_id}")
    
    # 通知审核员和管理员：有新的跳过申请
    try:
        pending_skips = db.query(Task).filter(Task.status == "skip_pending").count()
        from_name = getattr(current_user, 'real_name', None) or current_user.username
        content = f"{from_name} 提交了任务（{task.title}）的跳过申请，待审核"
        
        # 通知审核员
        logger.info(f"🔔 [TaskAPI] 广播跳过申请 -> reviewer, 待审核跳过: {pending_skips}")
        await ws_manager.broadcast_to_role(
            role="reviewer",
            message={
                "type": "skip_requested",
                "title": "有新的跳过申请",
                "content": content,
                "pending_skip": pending_skips,
                "task_id": task_id
            }
        )
        
        # 通知管理员
        logger.info(f"🔔 [TaskAPI] 广播跳过申请 -> admin, 待审核跳过: {pending_skips}")
        await ws_manager.broadcast_to_role(
            role="admin",
            message={
                "type": "skip_requested",
                "title": "有新的跳过申请",
                "content": content,
                "pending_skip": pending_skips,
                "task_id": task_id
            }
        )
    except Exception as _e:
        logger.warning(f"通知审核员和管理员跳过申请失败: {_e}")
    
    return {
        "success": True,
        "message": "跳过申请已提交，请等待审核",
        "task_id": task_id,
        "status": "skip_pending"
    }

@router.post("/{task_id}/review-skip")
async def review_skip_request(
    task_id: str,
    payload: TaskSkipReview,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("TaskReview"))
):
    """审核跳过申请（需菜单权限 TaskReview）"""
    logger.info(f"📋 [TaskAPI] 审核跳过申请: {task_id}, 审核人: {current_user.username}, 结果: {'同意' if payload.approved else '拒绝'}")
    
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        logger.warning(f"❌ [TaskAPI] 任务不存在: {task_id}")
        raise HTTPException(status_code=404, detail="任务不存在")
    
    # 验证状态：只有跳过申请状态的任务才能被审核
    if task.status != "skip_pending":
        logger.warning(f"❌ [TaskAPI] 任务状态不允许审核跳过申请: {task_id}, 当前状态: {task.status}")
        raise HTTPException(status_code=400, detail="任务状态不允许审核跳过申请")
    
    # 记录原始分配人
    assigned_user_id = task.assigned_to
    
    # 更新任务状态
    reviewed_time = utc_now()
    task.skip_reviewed_at = reviewed_time
    task.skip_reviewed_by = current_user.id
    task.skip_review_comment = payload.comment
    
    if payload.approved:
        # 同意跳过：将任务标记为已跳过
        task.status = "skipped"
        task.skipped_at = reviewed_time
        task.skip_reason = task.skip_request_reason  # 复制申请原因到跳过原因
        task.skip_images = task.skip_request_images  # 复制申请截图到跳过截图
        
        # 清除分配信息
        task.assigned_to = None
        task.assigned_to_name = None
        
        action_type = "skip_approved"
        message = "跳过申请已同意，任务已跳过"
    else:
        # 拒绝跳过：将任务状态恢复为进行中
        task.status = "in_progress"
        action_type = "skip_rejected"
        message = "跳过申请已拒绝，任务恢复为进行中"
    
    # 记录时间轴
    events = list(task.timeline or [])
    events.append({
        "type": action_type,
        "time": reviewed_time.isoformat(),
        "user_id": current_user.id,
        "user_name": getattr(current_user, 'real_name', None) or current_user.username,
        "comment": payload.comment,
        "approved": payload.approved
    })
    task.timeline = events
    flag_modified(task, 'timeline')
    
    # ✅ 更新项目统计字段
    if task.project_id:
        try:
            update_project_stats(db, task.project_id)
        except Exception as e:
            logger.error(f"❌ [TaskAPI] 更新项目统计失败: {e}")
    
    db.commit()
    
    # ✅ 清除缓存（重要：三重清除策略）
    # 1. 清除标注员的任务缓存
    if assigned_user_id:
        cache_service.invalidate_tasks_cache(task.project_id, assigned_user_id)
    # 2. 清除项目的所有任务缓存
    cache_service.invalidate_tasks_cache(task.project_id)
    # 3. 清除跨项目的任务缓存（审核页面可能查看所有项目）
    cache_service.invalidate_tasks_cache()
    # 4. 清除任务详情
    cache_service.invalidate_task_detail_cache(task_id)
    cache_service.invalidate_project_detail_cache(task.project_id)
    
    logger.info(f"✅ [TaskAPI] 跳过申请审核完成: {task_id}, 结果: {message}")
    logger.info(f"✅ [TaskAPI] 跳过审核缓存已清除: project={task.project_id}, user={assigned_user_id}, 所有视图已刷新")
    audit_logger.info(f"user_id={current_user.id} action=review_skip_request task_id={task_id} result={task.status}")

    # 通知标注员：跳过申请审核结果
    try:
        if assigned_user_id:
            result_type = "skip_approved" if payload.approved else "skip_rejected"
            result_content = (
                f"你的跳过申请已同意，任务（{task.title}）已标记为已跳过"
                if payload.approved else f"你的跳过申请被拒绝，任务（{task.title}）已恢复为进行中"
            )
            logger.info(f"🔔 [TaskAPI] 准备向标注员 {assigned_user_id} 发送跳过审核结果: {result_type} ({task.title})")
            await ws_manager.send_to_user_id(assigned_user_id, {
                "type": result_type,
                "title": "跳过申请审核结果",
                "content": result_content,
                "task_id": task_id
            })
            logger.info(f"🔔 [TaskAPI] 已向标注员 {assigned_user_id} 发送跳过审核结果: {result_type} ({task.title})")
    except Exception as _e:
        logger.warning(f"通知标注员跳过审核结果失败: {_e}")
    
    return {
        "success": True,
        "message": message,
        "task_id": task_id,
        "status": task.status,
        "assigned_to": assigned_user_id if not payload.approved else None
    }